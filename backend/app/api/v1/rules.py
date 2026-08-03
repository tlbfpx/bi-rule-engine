"""规则管理 API"""
from io import BytesIO

from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.exceptions import NotFoundException
from app.db import get_db
from app.models.rule import Rule
from app.models.rule_set import RuleSet
from app.schemas.rule import RuleCreate, RuleUpdate, RuleTestRequest, BatchPriorityUpdate, RuleOut
from app.schemas.common import Page
from app.services.rule_service import test_rule as run_rule_test
from app.services.rule_validator import validate_rule_config

router = APIRouter()


@router.get("", response_model=Page[RuleOut])
async def list_rules(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    field_name: str | None = None,
    rule_type: str | None = None,
    enabled: bool | None = None,
    rule_set_id: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(Rule)
    count_query = select(func.count(Rule.id))
    if field_name:
        query = query.where(Rule.field_name.ilike(f"%{field_name}%"))
        count_query = count_query.where(Rule.field_name.ilike(f"%{field_name}%"))
    if rule_type:
        query = query.where(Rule.rule_type == rule_type)
        count_query = count_query.where(Rule.rule_type == rule_type)
    if enabled is not None:
        query = query.where(Rule.enabled == enabled)
        count_query = count_query.where(Rule.enabled == enabled)
    if rule_set_id:
        query = query.where(Rule.rule_set_id == rule_set_id)
        count_query = count_query.where(Rule.rule_set_id == rule_set_id)
    total = (await db.execute(count_query)).scalar()
    query = query.order_by(Rule.priority.asc(), Rule.updated_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    rules = result.scalars().all()

    # 批量加载规则集名称
    rule_set_ids = {r.rule_set_id for r in rules if r.rule_set_id}
    rule_set_names = {}
    if rule_set_ids:
        from app.models.rule_set import RuleSet
        rs_result = await db.execute(select(RuleSet).where(RuleSet.id.in_(rule_set_ids)))
        for rs in rs_result.scalars().all():
            rule_set_names[rs.id] = rs.name

    # 把规则集名作为动态属性附在 ORM 实例上，由 response_model(RuleOut) 序列化
    for r in rules:
        r.rule_set_name = rule_set_names.get(r.rule_set_id)
        r.config_errors = validate_rule_config(r.rule_type, r.config or {})

    return {"items": rules, "total": total, "page": page, "page_size": page_size}


@router.post("", status_code=201, response_model=RuleOut)
async def create_rule(body: RuleCreate, db: AsyncSession = Depends(get_db)):
    # 校验配置完整性
    errors = validate_rule_config(body.rule_type, body.config.model_dump())
    if errors:
        from app.core.exceptions import BizException
        raise BizException(
            detail="规则配置不完整，请检查条件设置",
            data={"config_errors": errors},
        )

    rule = Rule(
        rule_set_id=body.rule_set_id,
        field_name=body.field_name, field_label=body.field_label,
        rule_type=body.rule_type, priority=body.priority, enabled=body.enabled,
        config=body.config.model_dump(), lookup_table_id=body.lookup_table_id,
        depends_on=body.depends_on or [], description=body.description,
    )
    db.add(rule)
    await db.flush()
    await db.refresh(rule)
    logger.info(f"创建规则: {rule.field_name} ({rule.rule_type})")
    return rule


@router.put("/batch-priority")
async def batch_update_priority(body: BatchPriorityUpdate, db: AsyncSession = Depends(get_db)):
    for item in body.items:
        result = await db.execute(select(Rule).where(Rule.id == item.id))
        rule = result.scalar_one_or_none()
        if rule:
            rule.priority = item.priority
    await db.flush()
    return {"message": "ok", "count": len(body.items)}


_OP_LABELS = {
    "eq": "=", "ne": "≠", "gt": ">", "gte": "≥", "lt": "<", "lte": "≤",
    "contains": "包含", "not_contains": "不包含",
    "starts_with": "开头是", "ends_with": "结尾是",
    "in": "属于", "not_in": "不属于",
    "is_null": "为空", "is_not_null": "不为空",
    "matches": "匹配", "not_matches": "不匹配",
}


def _format_result_val(result_val, result_type="constant"):
    """安全格式化结果值，正确处理 0、False、空字符串"""
    if result_type == "field_value":
        prefix = "取字段"
    else:
        prefix = "常量"
    if result_val is None or result_val == "":
        return f"{prefix}（空）"
    return f"{prefix} {result_val}"


def _format_default_result(config: dict) -> str:
    """格式化默认值，None/空字符串表示未设置"""
    default = config.get("default_result")
    if default is None or default == "":
        return ""
    return f"（默认: {default}）"


def _flatten_config(rule: Rule) -> str:
    """将规则配置展平为可读描述文本"""
    config = rule.config or {}
    rt = rule.rule_type
    if rt == "mapping":
        groups = config.get("conditions", [])
        total_rows = sum(len(g.get("rows", [])) for g in groups)
        parts = [f"条件组 {len(groups)} 组, 条件行 {total_rows} 行"]
        for g in groups:
            priority = g.get("priority", "?")
            logic = g.get("logic", "AND")
            result_type = g.get("result_type", "constant")
            result_val = g.get("result_value")

            # 构建条件行描述
            rows = g.get("rows", [])
            cond_segments = []
            for row in rows:
                field = row.get("field", "?")
                op = row.get("operator", "?")
                val = row.get("value")
                op_label = _OP_LABELS.get(op, op)

                if op in ("is_null", "is_not_null"):
                    cond_segments.append(f"{field} {op_label}")
                elif val is not None:
                    cond_segments.append(f"{field} {op_label} {val}")
                else:
                    cond_segments.append(f"{field} {op_label}")

            # 用"且"/"或"连接多个条件
            connector = " 且 " if logic == "AND" else " 或 "
            condition_str = connector.join(cond_segments) if cond_segments else "无条件"

            result_str = _format_result_val(result_val, result_type)
            parts.append(f"[P{priority} {logic}] 当 {condition_str} → {result_str}")

        # 追加默认值
        default_str = _format_default_result(config)
        if default_str:
            parts.append(default_str)
        return "\n".join(parts)
    elif rt == "cleaning":
        steps = config.get("cleaning_steps", [])
        actions = [s.get("action", "?") for s in steps]
        line = f"清洗步骤 {len(steps)} 步: {', '.join(actions)}"
        default_str = _format_default_result(config)
        return f"{line}\n{default_str}" if default_str else line
    elif rt == "lookup":
        tid = config.get("lookup_table_id") or rule.lookup_table_id or "-"
        kf = config.get("lookup_key_field", "-")
        vf = config.get("lookup_value_field", "-")
        line = f"查表: {tid} (key={kf}, value={vf})"
        default_str = _format_default_result(config)
        return f"{line}\n{default_str}" if default_str else line
    elif rt == "computed":
        expr = config.get("formula_expression", "")
        line = f"公式: {expr}" if expr else "公式: (空)"
        default_str = _format_default_result(config)
        return f"{line}\n{default_str}" if default_str else line
    return str(config)


@router.get("/export")
async def export_rules(
    rule_set_id: str = Query(..., min_length=1, description="规则集 ID"),
    db: AsyncSession = Depends(get_db),
):
    """导出指定规则集的所有规则为 Excel 文件"""
    # 查询规则集名称
    rs_result = await db.execute(select(RuleSet).where(RuleSet.id == rule_set_id))
    rule_set = rs_result.scalar_one_or_none()
    rs_name = rule_set.name if rule_set else rule_set_id

    # 查询所有规则
    result = await db.execute(
        select(Rule).where(Rule.rule_set_id == rule_set_id).order_by(Rule.priority.asc())
    )
    rules = result.scalars().all()

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "规则配置"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头
    headers = [
        "序号", "字段名", "字段标签", "规则类型", "优先级", "启用",
        "依赖规则", "配置说明", "描述", "创建时间", "更新时间",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    rule_type_labels = {
        "mapping": "映射", "cleaning": "清洗", "lookup": "查表", "computed": "计算",
    }
    for idx, rule in enumerate(rules, 1):
        row_data = [
            idx,
            rule.field_name,
            rule.field_label or "",
            rule_type_labels.get(rule.rule_type, rule.rule_type),
            rule.priority,
            "是" if rule.enabled else "否",
            ", ".join(rule.depends_on) if rule.depends_on else "",
            _flatten_config(rule),
            rule.description or "",
            rule.created_at.strftime("%Y-%m-%d %H:%M") if rule.created_at else "",
            rule.updated_at.strftime("%Y-%m-%d %H:%M") if rule.updated_at else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 列宽
    col_widths = [6, 18, 14, 10, 8, 6, 20, 42, 24, 18, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 输出
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"rules_{rs_name}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/{rule_id}", response_model=RuleOut)
async def get_rule(rule_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Rule).where(Rule.id == rule_id))
    rule = result.scalar_one_or_none()
    if not rule:
        raise NotFoundException(detail="规则不存在")
    return rule


@router.put("/{rule_id}", response_model=RuleOut)
async def update_rule(rule_id: str, body: RuleUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Rule).where(Rule.id == rule_id))
    rule = result.scalar_one_or_none()
    if not rule:
        raise NotFoundException(detail="规则不存在")
    update_data = body.model_dump(exclude_unset=True)

    # 校验配置完整性（仅在更新 config 时）
    new_rule_type = update_data.get("rule_type", rule.rule_type)
    if "config" in update_data and update_data["config"] is not None:
        if hasattr(update_data["config"], "model_dump"):
            update_data["config"] = update_data["config"].model_dump()
        errors = validate_rule_config(new_rule_type, update_data["config"])
        if errors:
            from app.core.exceptions import BizException
            raise BizException(
                detail="规则配置不完整，请检查条件设置",
                data={"config_errors": errors},
            )

    for key, value in update_data.items():
        setattr(rule, key, value)
    await db.flush()
    await db.refresh(rule)
    logger.info(f"更新规则: {rule.field_name}")
    return rule


@router.delete("/{rule_id}", status_code=204)
async def delete_rule(rule_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Rule).where(Rule.id == rule_id))
    rule = result.scalar_one_or_none()
    if not rule:
        raise NotFoundException(detail="规则不存在")
    await db.delete(rule)
    logger.info(f"删除规则: {rule.field_name}")


@router.post("/{rule_id}/test")
async def test_rule(rule_id: str, body: RuleTestRequest, db: AsyncSession = Depends(get_db)):
    """试跑单条规则（业务逻辑见 services.rule_service）。"""
    return await run_rule_test(db, rule_id, body.test_rows)
